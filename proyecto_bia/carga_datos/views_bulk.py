# carga_datos/views_bulk.py
import io
import uuid
import math
import hashlib
import datetime
from typing import Dict, Any, List, Tuple

import pandas as pd
from django.db import transaction, models
from django.utils import timezone
from django.apps import apps
from django.http import HttpResponse

from django.db.models import Max, BigIntegerField, Q
from django.db.models.functions import Cast

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers as xl_numbers
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

from carga_datos.models import (
    BulkJob,
    StagingBulkChange,
    AuditLog,
    BaseDeDatosBia,
    BusinessKeyCounter,
)
from carga_datos.utils_preview import render_preview_table

# 🚦 permisos de negocio
from carga_datos.permissions import CanBulkModify, IsAdminOrSuperuser


# =========================
# Configuración del módulo
# =========================
BUSINESS_KEY_FIELD = "id_pago_unico"
ALLOW_INSERTS = True
ALLOW_DELETES = True
ACCEPTED_OPS = {"UPDATE", "INSERT", "DELETE", "NOCHANGE"}

# Campo FK a resolver por nombre o id
ENTIDAD_FIELD = "entidad"
ENTIDAD_APP_LABEL = "certificado_ldd"
ENTIDAD_MODEL_NAME = "Entidad"

# Columnas NO editables (además de 'id')
# ⚠️ dni y cuit SON editables por pedido
NON_EDITABLE_FIELDS = {BUSINESS_KEY_FIELD, "fecha_apertura"}

# Requeridos para INSERT (tras normalizar)
REQUIRED_FOR_INSERT = {"dni", "cuit", "nombre_apellido"}

# Estados / subestados
ALLOWED_ESTADOS = {
    "NUEVO", "PENDIENTE", "CONTACTADO", "ACTIVO",
    "CANCELADO", "INCOBRABLE", "EN_GESTION", "CON DEUDA",
    "AGENCIA EXTERNA",
}
ALLOWED_SUBESTADOS_BY_ESTADO = {
    "NUEVO": {"SIN_CONTACTO", "ASIGNADO", "AGENCIA EXTERNA"},
    "PENDIENTE": {"A_VERIFICAR", "EN_ANALISIS", "AGENCIA EXTERNA"},
    "CONTACTADO": {"EN_NEGOCIACION", "SEGUIMIENTO", "AGENCIA EXTERNA"},
    "ACTIVO": {"PLAN_VIGENTE", "PAGO_PARCIAL", "AGENCIA EXTERNA"},
    "CANCELADO": {"PAGO_TOTAL", "ACUERDO_CUMPLIDO", "AGENCIA EXTERNA"},
    "INCOBRABLE": {"SIN_DATOS", "DEVUELTO", "AGENCIA EXTERNA"},
    "EN_GESTION": {"DERIVADO", "REAGENDA", "AGENCIA EXTERNA"},
    "CON DEUDA": {"VENCIDO", "MORA", "AGENCIA EXTERNA"},
    "AGENCIA EXTERNA": {"DERIVADO"},
}
ACTIVE_ESTADOS_SET = {"CANCELADO", "CON DEUDA"}
MAX_ACTIVOS_POR_DNI_ENTIDAD = 1

# Validaciones soft
PHONE_FIELDS = {"tel1", "tel2", "tel3", "tel4", "tel5"}
EMAIL_FIELDS = {"mail1", "mail2", "mail3"}
DEC_POSITIVE_FIELDS = {
    "saldo_capital", "saldo_exigible", "interes_total", "saldo_actualizado",
    "anticipo", "importe", "total_plan", "pago_acumulado",
}
DEC_RATE_FIELDS = {"interes_diario"}  # 0..1

# Validación estricta de CUIT (desactivada por defecto)
ENFORCE_CUIT_CHECKSUM = False


# ============ Utilidades / normalizadores ============


def _sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _is_nan(x: Any) -> bool:
    try:
        return isinstance(x, float) and math.isnan(x)
    except Exception:
        return False


def _normalize_val(v: Any) -> Any:
    """NaN->None; str->trim (''->None); fechas->YYYY-MM-DD; demás: igual."""
    if _is_nan(v):
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s != "" else None
    if isinstance(v, (pd.Timestamp, datetime.datetime, datetime.date)):
        try:
            d = v.date() if isinstance(v, (pd.Timestamp, datetime.datetime)) else v
            return d.isoformat()
        except Exception:
            return str(v)
    return v


def _normalize_business_key(v: Any) -> str | None:
    v = _normalize_val(v)
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v)


def _model_concrete_fields(model_cls) -> Dict[str, models.Field]:
    """{campo: Field} concretos (sin M2M/reverse)."""
    res = {}
    for f in model_cls._meta.get_fields():
        if isinstance(f, models.Field) and not (f.many_to_many or f.one_to_many):
            res[f.name] = f
    return res


def _get_entidad_model():
    return apps.get_model(ENTIDAD_APP_LABEL, ENTIDAD_MODEL_NAME)


def _coerce_entidad_value(raw: Any) -> Tuple[Any, str]:
    """Resolver 'entidad' por id o por nombre (iexact)."""
    v = _normalize_val(raw)
    if v is None:
        return None, ""
    Entidad = _get_entidad_model()
    s = str(v)
    if s.isdigit():
        pk = int(s)
        exists = Entidad.objects.filter(pk=pk).only("id").exists()
        return (pk if exists else None), ("" if exists else "entidad: id inexistente")
    obj = Entidad.objects.filter(nombre__iexact=s).only("id").first()
    if obj:
        return obj.pk, ""
    return None, "entidad: no encontrada por nombre"


def _coerce_to_field(field: models.Field, value: Any) -> Tuple[Any, str]:
    if field.name == ENTIDAD_FIELD and isinstance(field, models.ForeignKey):
        return _coerce_entidad_value(value)
    try:
        return field.to_python(value), ""
    except Exception as e:
        return None, f"{field.name}: {e}"


# ---- Normalización y validación suave de DNI/CUIT ----


def _only_digits(val: Any) -> str | None:
    """2054685741.0 -> '2054685741'; elimina todo salvo dígitos."""
    if val is None:
        return None
    if isinstance(val, float):
        try:
            if val.is_integer():
                return str(int(val))
            return ''.join(ch for ch in str(val) if ch.isdigit())
        except Exception:
            pass
    s = str(val)
    digits = ''.join(ch for ch in s if ch.isdigit())
    return digits or None


def _normalize_dni(val: Any) -> str | None:
    return _only_digits(val)


def _cuit_is_valid(cuit: str) -> bool:
    if not cuit or not cuit.isdigit() or len(cuit) != 11:
        return False
    weights = [5, 4, 3, 2, 7, 6, 5, 4, 3, 2]
    nums = [int(x) for x in cuit]
    s = sum(d * w for d, w in zip(nums[:10], weights))
    dv = 11 - (s % 11)
    if dv == 11:
        dv = 0
    elif dv == 10:
        dv = 9
    return dv == nums[10]


def _normalize_cuit(val: Any) -> str | None:
    digits = _only_digits(val)
    if not digits:
        return None
    if ENFORCE_CUIT_CHECKSUM and len(digits) == 11:
        if not _cuit_is_valid(digits):
            return None
    return digits


# ======= Helpers de autoincremento seguro =======


def _allocate_sequential_ids_from_db_max(n: int) -> list[str]:
    """Toma el mayor id en DB y el contador para asignar n ids consecutivos, de forma atómica."""
    if n <= 0:
        return []
    with transaction.atomic():
        counter, _ = BusinessKeyCounter.objects.select_for_update().get_or_create(
            name='id_pago_unico', defaults={'last_value': 0}
        )
        db_max = BaseDeDatosBia.objects.aggregate(
            m=Max(Cast(BUSINESS_KEY_FIELD, BigIntegerField()))
        )["m"] or 0
        last = max(int(counter.last_value or 0), int(db_max))
        start = last + 1
        counter.last_value = last + n
        counter.save(update_fields=['last_value', 'updated_at'])
        return [str(i) for i in range(start, start + n)]


# ======= Validaciones de negocio =======


def _email_is_valid(val: str) -> bool:
    if not val:
        return True
    try:
        validate_email(val)
        return True
    except ValidationError:
        return False


def _phone_is_valid(val: str) -> bool:
    if not val:
        return True
    return str(val).isdigit()


def _dates_are_valid(fecha_apertura: Any, fecha_deuda: Any, today: datetime.date) -> Tuple[bool, str]:
    """fecha_deuda < fecha_apertura (si ambas); ninguna > hoy."""
    try:
        fa = datetime.date.fromisoformat(str(fecha_apertura)) if fecha_apertura else None
        fd = datetime.date.fromisoformat(str(fecha_deuda)) if fecha_deuda else None
    except Exception:
        return False, "Fechas con formato inválido (use YYYY-MM-DD)."
    if fa and fa > today:
        return False, "fecha_apertura no puede ser posterior a hoy."
    if fd and fd > today:
        return False, "fecha_deuda no puede ser posterior a hoy."
    if fa and fd and not (fd < fa):
        return False, "fecha_deuda debe ser anterior a fecha_apertura."
    return True, ""


def _money_is_valid(payload: dict) -> Tuple[bool, str]:
    for f in DEC_POSITIVE_FIELDS:
        if f in payload and payload[f] is not None:
            try:
                if float(payload[f]) < 0:
                    return False, f"'{f}' no puede ser negativo."
            except Exception:
                return False, f"'{f}' debe ser numérico."
    for f in DEC_RATE_FIELDS:
        if f in payload and payload[f] is not None:
            try:
                v = float(payload[f])
                if v < 0 or v > 1:
                    return False, f"'{f}' debe estar entre 0 y 1."
            except Exception:
                return False, f"'{f}' debe ser numérico."
    try:
        if payload.get("saldo_actualizado") is not None and payload.get("saldo_exigible") is not None:
            if float(payload["saldo_actualizado"]) < float(payload["saldo_exigible"]):
                return False, "saldo_actualizado debe ser ≥ saldo_exigible."
    except Exception:
        return False, "Campos de saldo deben ser numéricos."
    return True, ""


def _normalize_estado(val: Any) -> str:
    return str(val).strip().upper() if val is not None else ""


def _estado_is_valid(estado: str, sub_estado: str) -> Tuple[bool, str]:
    est = _normalize_estado(estado) if estado else ""
    sub = _normalize_estado(sub_estado) if sub_estado else ""
    if est and est not in ALLOWED_ESTADOS:
        return False, f"Estado inválido: '{estado}'."
    if sub:
        if sub == "AGENCIA EXTERNA":
            return True, ""
        if not est:
            return False, "Sub-estado requiere un estado base."
        allowed = ALLOWED_SUBESTADOS_BY_ESTADO.get(est, set())
        if allowed and sub not in allowed:
            return False, f"Sub-estado '{sub_estado}' no permitido para estado '{estado}'."
    return True, ""


def _is_active_estado(estado: str) -> bool:
    return _normalize_estado(estado) in ACTIVE_ESTADOS_SET


def _soft_uniqueness_ok(dni: str, entidad_id: int, estado: str) -> Tuple[bool, str]:
    if not dni or not entidad_id:
        return True, ""
    if not _is_active_estado(estado):
        return True, ""
    qs = BaseDeDatosBia.objects.filter(
        dni=dni, entidad_id=entidad_id
    ).filter(Q(estado__iexact="CANCELADO") | Q(estado__iexact="CON DEUDA"))
    cnt = qs.count()
    if cnt >= MAX_ACTIVOS_POR_DNI_ENTIDAD:
        return False, f"Regla de negocio: ya existen {cnt} registro(s) activo(s) para este DNI+Entidad (estados: CANCELADO / CON DEUDA)."
    return True, ""


# =========== EXPORT .XLSX ===========


@api_view(["GET"])
@permission_classes([IsAuthenticated, CanBulkModify])
def bulk_export_xlsx(request):
    """
    Descarga la base real (db_bia) en .xlsx:
      - Todas las columnas (sin 'id')
      - Agrega '__op' vacía
      - ❗ Formatea DNI y CUIT como TEXTO para evitar .0/notación científica en Excel
    """
    model_fields = [f.name for f in BaseDeDatosBia._meta.fields if f.name != "id"]
    base_cols = [BUSINESS_KEY_FIELD] + [c for c in model_fields if c != BUSINESS_KEY_FIELD]
    export_cols = [BUSINESS_KEY_FIELD, "__op"] + [c for c in base_cols if c != BUSINESS_KEY_FIELD]

    qs = BaseDeDatosBia.objects.all().order_by("id")

    dni = (request.GET.get("dni") or "").strip()
    if dni:
        qs = qs.filter(dni=dni)

    idp_csv = (request.GET.get(f"{BUSINESS_KEY_FIELD}__in") or "").strip()
    if idp_csv:
        ids = [s.strip() for s in idp_csv.split(",") if s.strip()]
        qs = qs.filter(**{f"{BUSINESS_KEY_FIELD}__in": ids})

    try:
        limit = int(request.GET.get("limit", "0"))
    except Exception:
        limit = 0
    if limit > 0:
        qs = qs[:limit]

    rows = list(qs.values(*base_cols))
    norm_rows = []
    for r in rows:
        norm = {k: _normalize_val(v) for k, v in r.items()}
        norm["__op"] = ""
        # Normalizar dni/cuit para exportar como texto limpio
        if "dni" in norm and norm["dni"] is not None:
            norm["dni"] = _normalize_dni(norm["dni"]) or norm["dni"]
        if "cuit" in norm and norm["cuit"] is not None:
            norm["cuit"] = _normalize_cuit(norm["cuit"]) or norm["cuit"]
        norm_rows.append(norm)

    df = pd.DataFrame(norm_rows, columns=export_cols)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Datos")
        ws = writer.sheets["Datos"]
        ws.freeze_panes = "A2"
        # Ancho de columnas
        for i, col in enumerate(export_cols, start=1):
            width = max(10, min(40, len(str(col)) + 2))
            ws.column_dimensions[get_column_letter(i)].width = width

        # ❗ Forzar formato TEXTO en columnas DNI y CUIT
        def _col_index(col_name: str) -> int | None:
            try:
                return export_cols.index(col_name) + 1  # 1-based
            except ValueError:
                return None

        dni_col = _col_index("dni")
        cuit_col = _col_index("cuit")
        for col_idx in filter(None, [dni_col, cuit_col]):
            col_letter = get_column_letter(col_idx)
            # desde fila 2 (fila 1 es header)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row_idx}"]
                cell.number_format = "@"  # texto

        # Hoja Guía
        guia_lines = [
            ["Instrucciones"],
            [f"- En UPDATE, NO modificar '{BUSINESS_KEY_FIELD}' ni 'fecha_apertura'."],
            ["- '__op': UPDATE (o vacío), INSERT, DELETE."],
            ["- INSERT: 'id_pago_unico' puede ir vacío (se autogenera); 'fecha_apertura' se completa si falta."],
            ["- Fechas: YYYY-MM-DD; 'fecha_deuda' debe ser anterior a 'fecha_apertura'."],
            ["- 'entidad': ID numérico o nombre (case-insensitive)."],
            ["- Estados/sub-estados permiten 'Agencia Externa'."],
            ["- DNI/CUIT exportados como TEXTO para evitar .0 en Excel."],
        ]
        pd.DataFrame([{"Guía": x[0]} for x in guia_lines]).to_excel(writer, index=False, sheet_name="Guía")
        writer.sheets["Guía"].column_dimensions["A"].width = 110

    buffer.seek(0)
    ts = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    filename = f"db_bia_export_{ts}.xlsx"

    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# =========== VALIDATE ===========


@api_view(["POST"])
@permission_classes([IsAuthenticated, CanBulkModify])
def bulk_validate(request):
    """
    Sube un archivo (CSV/XLS/XLSX) y prepara la vista previa de cambios.
    Reglas clave:
    - INSERT: dni/cuit/nombre_apellido obligatorios (tras normalizar); fecha_apertura auto si falta;
              fecha_deuda < fecha_apertura; unicidad blanda (CANCELADO/CON DEUDA).
    - UPDATE: ignora celdas vacías (incluye NO-EDITABLES vacías). NO-EDITABLES con valor => error.
    - DELETE: basta id_pago_unico + __op=DELETE (otras columnas se ignoran).
    """
    f = request.FILES.get("archivo")
    if not f:
        return Response({"errors": ["Archivo requerido."]}, status=400)

    file_bytes = f.read()
    file_hash = _sha256_bytes(file_bytes)
    name = f.name.lower()

    # Leer en pandas
    if name.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(file_bytes))
    else:
        df = pd.read_csv(io.BytesIO(file_bytes))

    if df.empty:
        return Response({"errors": ["El archivo está vacío."]}, status=400)

    # Normalizar headers
    df.rename(columns={c: str(c).strip() for c in df.columns}, inplace=True)
    cols = [str(c).strip() for c in df.columns]

    if BUSINESS_KEY_FIELD not in cols and "business_key" not in cols:
        return Response({"errors": [f"Falta la columna clave '{BUSINESS_KEY_FIELD}'"]}, status=400)

    fields_map = _model_concrete_fields(BaseDeDatosBia)
    editable_cols = set(fields_map.keys()) - ({"id"} | NON_EDITABLE_FIELDS)

    job = BulkJob.objects.create(
        id=uuid.uuid4(),
        filename=f.name,
        file_hash=file_hash,
        created_by=request.user,
        status=BulkJob.Status.READY,
        summary={},
    )

    summary = {"total": int(len(df)), "ok": 0, "con_errores": 0, "updates": 0, "inserts": 0, "deletes": 0, "nochange": 0}
    preview_rows: List[Dict[str, Any]] = []
    today = timezone.localdate()

    with transaction.atomic():
        for _, s in df.iterrows():
            raw_row = {k: _normalize_val(v) for k, v in s.to_dict().items()}
            bkey = _normalize_business_key(raw_row.get(BUSINESS_KEY_FIELD) or raw_row.get("business_key"))
            op_in = str(raw_row.get("__op") or "").upper().strip()
            if op_in and op_in not in ACCEPTED_OPS:
                op_in = ""

            # DELETE: ignorar demás columnas
            if op_in == "DELETE":
                current = BaseDeDatosBia.objects.filter(**{f"{BUSINESS_KEY_FIELD}": bkey}).first()
                op = "DELETE" if current else "NOCHANGE"
                errors = []
                if not current:
                    errors.append("DELETE ignorado: la clave no existe en DB.")
                can_apply = (op == "DELETE") and len(errors) == 0 and ALLOW_DELETES
                if not ALLOW_DELETES:
                    errors.append("Borrado masivo deshabilitado por configuración.")
                    can_apply = False

                StagingBulkChange.objects.update_or_create(
                    job=job,
                    business_key=str(bkey) if bkey else "(sin_clave)",
                    defaults=dict(op=op, payload=raw_row, validation_errors=errors, can_apply=can_apply),
                )
                summary["deletes"] += 1 if op == "DELETE" else 0
                summary["nochange"] += 1 if op == "NOCHANGE" else 0
                if can_apply:
                    summary["ok"] += 1
                else:
                    summary["con_errores"] += 1
                preview_rows.append({BUSINESS_KEY_FIELD: str(bkey) if bkey else "(sin_clave)", "op": op, "errors": errors, "changes": {}})
                continue

            # Construir payload_clean (ignorando vacíos; no editables vacíos => ignorar)
            payload_clean = {}
            errors: List[str] = []
            for k, v in raw_row.items():
                if k in {BUSINESS_KEY_FIELD, "business_key", "__op"}:
                    continue
                if k not in fields_map:
                    if _normalize_val(v) is not None:
                        errors.append(f"Columna desconocida: {k}")
                    continue

                is_non_editable = k in NON_EDITABLE_FIELDS
                norm_v = _normalize_val(v)

                # NO-EDITABLE vacía => ignorar
                if is_non_editable and norm_v is None:
                    continue
                # NO-EDITABLE con valor => error
                if is_non_editable and norm_v is not None:
                    errors.append(f"Columna no editable: {k}")
                    continue

                if k not in editable_cols:
                    continue
                if norm_v is None:
                    continue

                # Normalización especial para dni/cuit
                if k == "dni":
                    nd = _normalize_dni(norm_v)
                    if nd:
                        payload_clean[k] = nd
                    continue
                if k == "cuit":
                    nc = _normalize_cuit(norm_v)
                    if nc:
                        payload_clean[k] = nc
                    continue

                coerced, err = _coerce_to_field(fields_map[k], norm_v)
                if err:
                    errors.append(err)
                else:
                    payload_clean[k] = coerced

            current = BaseDeDatosBia.objects.filter(**{f"{BUSINESS_KEY_FIELD}": bkey}).first() if bkey else None

            changes = {}
            if current:
                # UPDATE/NOCHANGE
                for k, newv in payload_clean.items():
                    oldv = getattr(current, f"{k}_id") if isinstance(fields_map[k], models.ForeignKey) else getattr(current, k, None)
                    cmp_old = None if oldv is None else str(oldv)
                    cmp_new = None if newv is None else (str(newv) if not hasattr(newv, "pk") else str(newv))
                    if cmp_old != cmp_new:
                        if k == ENTIDAD_FIELD:
                            old_display = None
                            if getattr(current, f"{k}_id") is not None:
                                old_ent = getattr(current, k)
                                old_display = f"{old_ent.id} · {getattr(old_ent, 'nombre', '')}"
                            new_display = None
                            if newv is not None:
                                Entidad = _get_entidad_model()
                                ent = Entidad.objects.filter(pk=newv).only("id", "nombre").first()
                                new_display = f"{ent.id} · {ent.nombre}" if ent else str(newv)
                            changes[k] = {"old": old_display, "new": new_display}
                        else:
                            changes[k] = {"old": getattr(current, k, None), "new": newv}
                op = "UPDATE" if changes else "NOCHANGE"

                if op == "UPDATE" and payload_clean:
                    # Emails/Teléfonos (suaves)
                    for ef in EMAIL_FIELDS:
                        if ef in payload_clean and not _email_is_valid(str(payload_clean[ef])):
                            errors.append(f"Email inválido en {ef}.")
                    for pf in PHONE_FIELDS:
                        if pf in payload_clean and not _phone_is_valid(str(payload_clean[pf])):
                            errors.append(f"Teléfono inválido (solo dígitos) en {pf}.")

                    # Fechas
                    fa = payload_clean.get("fecha_apertura") or getattr(current, "fecha_apertura", None)
                    fd = payload_clean.get("fecha_deuda") or getattr(current, "fecha_deuda", None)
                    okd, msgd = _dates_are_valid(fa, fd, today)
                    if not okd:
                        errors.append(msgd)

                    # Montos
                    okm, msgm = _money_is_valid(payload_clean)
                    if not okm:
                        errors.append(msgm)

                    # Estado/Subestado
                    estado_val = _normalize_estado(payload_clean.get("estado") or getattr(current, "estado", None))
                    sub_estado_val = _normalize_estado(payload_clean.get("sub_estado") or getattr(current, "sub_estado", None))
                    oke, msge = _estado_is_valid(estado_val, sub_estado_val)
                    if not oke:
                        errors.append(msge)

                    # Unicidad blanda (solo si estado activo)
                    entidad_id = payload_clean.get("entidad") or getattr(current, "entidad_id", None)
                    dni_val = payload_clean.get("dni") or getattr(current, "dni", None)
                    if entidad_id and dni_val and _is_active_estado(estado_val):
                        cnt = BaseDeDatosBia.objects.filter(
                            dni=dni_val, entidad_id=entidad_id
                        ).filter(Q(estado__iexact="CANCELADO") | Q(estado__iexact="CON DEUDA")).exclude(pk=current.pk).count()
                        if cnt >= MAX_ACTIVOS_POR_DNI_ENTIDAD:
                            errors.append(f"Regla de negocio: ya existen {cnt} registro(s) activo(s) para este DNI+Entidad (estados: CANCELADO / CON DEUDA).")

            else:
                # INSERT
                op = "INSERT"

                # Requeridos (tras normalizar)
                dni_norm = _normalize_dni(raw_row.get("dni"))
                cuit_norm = _normalize_cuit(raw_row.get("cuit"))
                nombre_ok = bool(_normalize_val(raw_row.get("nombre_apellido")))
                if not dni_norm:
                    errors.append("Falta DNI (no se pudo normalizar).")
                if not cuit_norm:
                    errors.append("Falta CUIT (no se pudo normalizar).")
                if not nombre_ok:
                    errors.append("Falta nombre_apellido.")

                if dni_norm:
                    payload_clean["dni"] = dni_norm
                if cuit_norm:
                    payload_clean["cuit"] = cuit_norm

                # Fechas: fecha_apertura hoy si falta; fecha_deuda < fecha_apertura
                fa = payload_clean.get("fecha_apertura") or _normalize_val(raw_row.get("fecha_apertura")) or today
                fd = payload_clean.get("fecha_deuda") or _normalize_val(raw_row.get("fecha_deuda"))
                okd, msgd = _dates_are_valid(fa, fd, today)
                if not okd:
                    errors.append(msgd)

                # Montos
                okm, msgm = _money_is_valid(payload_clean)
                if not okm:
                    errors.append(msgm)

                # Estado/Subestado
                estado_val = _normalize_estado(payload_clean.get("estado") or raw_row.get("estado"))
                sub_estado_val = _normalize_estado(payload_clean.get("sub_estado") or raw_row.get("sub_estado"))
                oke, msge = _estado_is_valid(estado_val, sub_estado_val)
                if not oke:
                    errors.append(msge)

                # Unicidad blanda si activo
                entidad_id = payload_clean.get("entidad")
                if entidad_id is None and _normalize_val(raw_row.get("entidad")):
                    coerced_ent, err_ent = _coerce_to_field(fields_map["entidad"], raw_row.get("entidad"))
                    if not err_ent:
                        entidad_id = coerced_ent
                        payload_clean["entidad"] = entidad_id
                dni_val = payload_clean.get("dni")
                if entidad_id and dni_val and _is_active_estado(estado_val):
                    okuniq, msguniq = _soft_uniqueness_ok(dni_val, entidad_id, estado_val)
                    if not okuniq:
                        errors.append(msguniq)

                # Preview de cambios (incluye (auto) para id)
                for k, v in payload_clean.items():
                    if k == ENTIDAD_FIELD and v is not None:
                        Entidad = _get_entidad_model()
                        ent = Entidad.objects.filter(pk=v).only("id", "nombre").first()
                        disp = f"{ent.id} · {ent.nombre}" if ent else str(v)
                        changes[k] = {"old": None, "new": disp}
                    else:
                        changes[k] = {"old": None, "new": v}
                if not bkey:
                    changes[BUSINESS_KEY_FIELD] = {"old": None, "new": "(auto)"}

            can_apply = (op in {"UPDATE", "INSERT"}) and len(errors) == 0
            StagingBulkChange.objects.update_or_create(
                job=job,
                business_key=str(bkey) if bkey else ("(auto)" if op == "INSERT" else "(sin_clave)"),
                defaults=dict(op=op, payload=raw_row, validation_errors=errors, can_apply=can_apply if op != "DELETE" else False),
            )

            if op == "UPDATE":
                summary["updates"] += 1
            elif op == "INSERT":
                summary["inserts"] += 1
            elif op == "NOCHANGE":
                summary["nochange"] += 1

            if op != "DELETE":
                if can_apply:
                    summary["ok"] += 1
                else:
                    summary["con_errores"] += 1

            preview_rows.append({
                BUSINESS_KEY_FIELD: str(bkey) if bkey else ("(auto)" if op == "INSERT" else "(sin_clave)"),
                "op": op,
                "errors": errors,
                "changes": changes,
            })

    job.summary = summary
    job.save(update_fields=["summary"])

    preview_html = render_preview_table(preview_rows, title="Vista previa — Modificación masiva")

    return Response({"success": True, "job_id": str(job.id), "preview": preview_html, "summary": summary})


# =========== COMMIT ===========


@api_view(["POST"])
@permission_classes([IsAuthenticated, IsAdminOrSuperuser])
def bulk_commit(request):
    """
    Aplica el job:
      - UPDATE/INSERT/DELETE
      - INSERT: autogenera id si falta (secuencial desde DB max) y fecha_apertura hoy si falta.
      - UPDATE: ignora vacíos (incluye NO-EDITABLES vacías).
      - DELETE: elimina por id_pago_unico.
      - Revalida reglas críticas.
      - AuditLog por campo.
    """
    job_id = request.data.get("job_id")
    if not job_id:
        return Response({"errors": ["'job_id' requerido."]}, status=400)

    job = BulkJob.objects.filter(id=job_id, status=BulkJob.Status.READY).first()
    if not job:
        return Response({"errors": ["Job inválido o ya procesado."]}, status=400)

    staging_qs = StagingBulkChange.objects.filter(job=job)
    if not staging_qs.exists():
        return Response({"errors": ["No hay staging para este job."]}, status=400)

    fields_map = _model_concrete_fields(BaseDeDatosBia)
    editable_cols = set(fields_map.keys()) - ({"id"} | NON_EDITABLE_FIELDS)

    rows = list(staging_qs.values("business_key", "op", "payload"))
    keys = [r["business_key"] for r in rows if r["business_key"] and r["business_key"] not in {"(auto)", "(sin_clave)"}]

    existentes = {getattr(o, BUSINESS_KEY_FIELD): o for o in BaseDeDatosBia.objects.filter(**{f"{BUSINESS_KEY_FIELD}__in": keys})}

    updates_instances = []
    inserts_instances = []
    deletes_keys = []
    changed_fields_union = set()
    pending_inserts_payloads = []
    today = timezone.localdate()

    with transaction.atomic():
        for r in rows:
            bkey = r["business_key"]
            op = (r["op"] or "").upper()
            raw = r["payload"] or {}

            # DELETE
            if op == "DELETE" and ALLOW_DELETES:
                if bkey in existentes:
                    deletes_keys.append(bkey)
                    AuditLog.objects.create(
                        table_name=BaseDeDatosBia._meta.db_table,
                        business_key=str(bkey),
                        field="*",
                        old_value="(row)",
                        new_value=None,
                        job=job,
                        action=AuditLog.Action.DELETE,
                        actor=request.user,
                    )
                continue

            # payload_clean (ignora vacíos; no aplica no-editables)
            payload_clean = {}
            for k, v in raw.items():
                if k in {BUSINESS_KEY_FIELD, "business_key", "__op"}:
                    continue
                if k not in fields_map:
                    continue
                norm_v = _normalize_val(v)

                if k in NON_EDITABLE_FIELDS:
                    # vacía -> ignorar; con valor -> ignorar (defensa)
                    continue
                if k not in editable_cols:
                    continue
                if norm_v is None:
                    continue

                if k == "dni":
                    nd = _normalize_dni(norm_v)
                    if nd:
                        payload_clean[k] = nd
                    continue
                if k == "cuit":
                    nc = _normalize_cuit(norm_v)
                    if nc:
                        payload_clean[k] = nc
                    continue

                coerced, err = _coerce_to_field(fields_map[k], norm_v)
                if err:
                    continue
                payload_clean[k] = coerced

            if op == "UPDATE":
                obj = existentes.get(bkey)
                if not obj:
                    continue

                # Revalidaciones
                fa = payload_clean.get("fecha_apertura") or getattr(obj, "fecha_apertura", None)
                fd = payload_clean.get("fecha_deuda") or getattr(obj, "fecha_deuda", None)
                okd, msgd = _dates_are_valid(fa, fd, today)
                if not okd:
                    return Response({"errors": [f"Fila {bkey}: {msgd}"]}, status=400)
                okm, msgm = _money_is_valid(payload_clean)
                if not okm:
                    return Response({"errors": [f"Fila {bkey}: {msgm}"]}, status=400)
                estado_val = _normalize_estado(payload_clean.get("estado") or getattr(obj, "estado", None))
                sub_estado_val = _normalize_estado(payload_clean.get("sub_estado") or getattr(obj, "sub_estado", None))
                oke, msge = _estado_is_valid(estado_val, sub_estado_val)
                if not oke:
                    return Response({"errors": [f"Fila {bkey}: {msge}"]}, status=400)
                for ef in EMAIL_FIELDS:
                    if ef in payload_clean and not _email_is_valid(str(payload_clean[ef])):
                        return Response({"errors": [f"Fila {bkey}: Email inválido en {ef}."]}, status=400)
                for pf in PHONE_FIELDS:
                    if pf in payload_clean and not _phone_is_valid(str(payload_clean[pf])):
                        return Response({"errors": [f"Fila {bkey}: Teléfono inválido en {pf} (solo dígitos)."]}, status=400)

                entidad_id = payload_clean.get("entidad") or getattr(obj, "entidad_id", None)
                dni_val = payload_clean.get("dni") or getattr(obj, "dni", None)
                if entidad_id and dni_val and _is_active_estado(estado_val):
                    cnt = BaseDeDatosBia.objects.filter(
                        dni=dni_val, entidad_id=entidad_id
                    ).filter(Q(estado__iexact="CANCELADO") | Q(estado__iexact="CON DEUDA")).exclude(pk=obj.pk).count()
                    if cnt >= MAX_ACTIVOS_POR_DNI_ENTIDAD:
                        return Response({"errors": [f"Fila {bkey}: Regla de negocio: ya existen {cnt} registro(s) activo(s) para este DNI+Entidad (estados: CANCELADO / CON DEUDA)."]}, status=400)

                # Aplicar
                local_changed = []
                for k, newv in payload_clean.items():
                    oldv = getattr(obj, f"{k}_id") if isinstance(fields_map[k], models.ForeignKey) else getattr(obj, k, None)
                    cmp_old = None if oldv is None else str(oldv)
                    cmp_new = None if newv is None else str(newv)
                    if cmp_old != cmp_new:
                        if isinstance(fields_map[k], models.ForeignKey):
                            setattr(obj, f"{k}_id", newv)
                        else:
                            setattr(obj, k, newv)
                        local_changed.append(k)
                        AuditLog.objects.create(
                            table_name=BaseDeDatosBia._meta.db_table,
                            business_key=str(bkey),
                            field=k,
                            old_value=str(oldv) if oldv is not None else None,
                            new_value=str(newv) if newv is not None else None,
                            job=job,
                            action=AuditLog.Action.UPDATE,
                            actor=request.user,
                        )
                if local_changed:
                    updates_instances.append(obj)
                    changed_fields_union.update(local_changed)

            elif op == "INSERT" and ALLOW_INSERTS:
                # Requeridos normalizados
                dni_norm = _normalize_dni(raw.get("dni"))
                cuit_norm = _normalize_cuit(raw.get("cuit"))
                nombre_ok = bool(_normalize_val(raw.get("nombre_apellido")))
                if not dni_norm or not cuit_norm or not nombre_ok:
                    return Response({"errors": [f"Fila {bkey or '(auto)'}: faltan dni/cuit/nombre_apellido (tras normalizar)."]}, status=400)
                payload_clean.setdefault("dni", dni_norm)
                payload_clean.setdefault("cuit", cuit_norm)

                # Fechas
                payload_clean.setdefault("fecha_apertura", today)
                okd, msgd = _dates_are_valid(payload_clean.get("fecha_apertura"), payload_clean.get("fecha_deuda"), today)
                if not okd:
                    return Response({"errors": [f"Fila {bkey or '(auto)'}: {msgd}"]}, status=400)

                # Montos
                okm, msgm = _money_is_valid(payload_clean)
                if not okm:
                    return Response({"errors": [f"Fila {bkey or '(auto)'}: {msgm}"]}, status=400)

                # Estado/Subestado
                estado_val = _normalize_estado(payload_clean.get("estado") or raw.get("estado"))
                sub_estado_val = _normalize_estado(payload_clean.get("sub_estado") or raw.get("sub_estado"))
                oke, msge = _estado_is_valid(estado_val, sub_estado_val)
                if not oke:
                    return Response({"errors": [f"Fila {bkey or '(auto)'}: {msge}"]}, status=400)

                # Unicidad blanda si activo
                entidad_id = payload_clean.get("entidad")
                if entidad_id is None and _normalize_val(raw.get("entidad")):
                    coerced_ent, err_ent = _coerce_to_field(fields_map["entidad"], raw.get("entidad"))
                    if not err_ent:
                        entidad_id = coerced_ent
                        payload_clean["entidad"] = entidad_id
                dni_val = payload_clean.get("dni")
                if entidad_id and dni_val and _is_active_estado(estado_val):
                    okuniq, msguniq = _soft_uniqueness_ok(dni_val, entidad_id, estado_val)
                    if not okuniq:
                        return Response({"errors": [f"Fila {bkey or '(auto)'}: {msguniq}"]}, status=400)

                pending_inserts_payloads.append({
                    "bkey": (bkey if (bkey and bkey not in {"(auto)", "(sin_clave)"}) else None),
                    "payload_clean": payload_clean.copy()
                })

        # Asignar ids automáticos
        need_auto = sum(1 for it in pending_inserts_payloads if not it["bkey"])
        auto_ids = _allocate_sequential_ids_from_db_max(need_auto) if need_auto else []
        auto_iter = iter(auto_ids)

        for it in pending_inserts_payloads:
            final_bkey = it["bkey"] or next(auto_iter)
            obj = BaseDeDatosBia(**{BUSINESS_KEY_FIELD: final_bkey})
            for k, v in it["payload_clean"].items():
                if isinstance(fields_map[k], models.ForeignKey):
                    setattr(obj, f"{k}_id", v)
                else:
                    setattr(obj, k, v)
            inserts_instances.append(obj)

            AuditLog.objects.create(
                table_name=BaseDeDatosBia._meta.db_table,
                business_key=str(final_bkey),
                field=BUSINESS_KEY_FIELD,
                old_value=None,
                new_value=str(final_bkey),
                job=job,
                action=AuditLog.Action.INSERT,
                actor=request.user,
            )
            for k, newv in it["payload_clean"].items():
                AuditLog.objects.create(
                    table_name=BaseDeDatosBia._meta.db_table,
                    business_key=str(final_bkey),
                    field=k,
                    old_value=None,
                    new_value=str(newv) if newv is not None else None,
                    job=job,
                    action=AuditLog.Action.INSERT,
                    actor=request.user,
                )

        # Persistencia
        if inserts_instances:
            BaseDeDatosBia.objects.bulk_create(inserts_instances, ignore_conflicts=True)
        if updates_instances and changed_fields_union:
            BaseDeDatosBia.objects.bulk_update(updates_instances, fields=list(changed_fields_union))
        if ALLOW_DELETES and deletes_keys:
            BaseDeDatosBia.objects.filter(**{f"{BUSINESS_KEY_FIELD}__in": deletes_keys}).delete()

        job.status = BulkJob.Status.COMMITTED
        job.committed_at = timezone.now()
        job.save(update_fields=["status", "committed_at"])

    return Response({
        "success": True,
        "inserted_count": len(inserts_instances),
        "updated_count": len(updates_instances),
        "deleted_count": len(deletes_keys) if ALLOW_DELETES else 0,
        "status": job.status,
    })
